import openpyxl

# ENDRE DENNE FOR Å ENDRE MND
month = "September"

book = openpyxl.load_workbook('input.xlsx', data_only=True)
sheet = book.active
rows = sheet.rows
inputData = []

## Read input data into inputData array
for row in rows:
    tempArray = []
    for cell in row:
        tempArray.append(cell.value)

    inputData.append(tempArray)

print("#####")
#print(inputData)

## Write inputData into tips.xlsx
wb = openpyxl.load_workbook(filename = 'tips.xlsx')
ws = wb[month]

i = 2
j = 1
employeeCount = 0
rowSum = 0
splitSum = 0

for excelRow in inputData:
    i += 1
    j = 1
    print(excelRow)
    #print(employeeCount)
    rowSum = excelRow[0]
    print("SUM = " + str(rowSum))
    for excelCell in excelRow:
        if (excelCell != None):
            employeeCount += 1
        j += 1
        ws.cell(row = i, column = j, value = excelCell)
    splitSum = rowSum / (employeeCount-1)
    print(str(employeeCount) +" employees")
    ws.cell(row = i, column = 16, value = int(splitSum))
    employeeCount = 0
            
        

wb.save('tips.xlsx')


#for row in rows.iter_rows(min_row = 1, min_col = 1, max_row = 3, max_col = 2):
#    for cell in row:
#        print(cell.value)
#    print()